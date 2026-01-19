
# This file contains the ExcelHandler class

# IMPORTS!
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from .gui import GUI

class ExcelHandler:
    """This class is reponsible for handling excel related functionality such as reading, appending, removing, copying etc.
    """

    __excel_filename: str = ''              # Name of the excel file that is being managed

    # constructor
    def __init__(self, filename: str) -> None:
        """Initialize an ExcelHandler instance.
        """
        # Validating filename
        assert type(filename) == str, "filename needs to be string"
        assert filename != "", "filename cannot be none"
        
        # initializing
        self.__excel_filename = filename

    # open_file
    def open_file(self, headers: list) -> openpyxl.Workbook:
        """Opens an Excel file, it also checks either the file aready exists or not,
        if not then it creates a new file with the specified headers.

        Args:
            filename (str): Name of the Excel file.
            headers (list): A list containing headers.

            Returns:
                _type_: None else openpyxl.Workbook instance.
        """
        wb = None
        # opening the excel file
        try:
            wb = openpyxl.load_workbook(filename=self.__excel_filename)
        except FileNotFoundError:       # if the file doesn't exist
            # creating an excel file
            wb = openpyxl.Workbook()
            # getting the active sheet
            ws = wb.active
            # appending headers to the top of the file
            ws.append(headers)
            # setting the title of the sheet
            ws.title = "Order_Details"
            # saving the workbook
            wb.save(self.__excel_filename)
        
        return wb
    
    # write
    def write(self, worksheet: Worksheet, data):
        """Writes order details on the excel file.
        """
        # need to check for duplicate orders before adding
        duplicate_orders = []
        # fetch existing orders from excel
        existing_orders = set()
        for row in worksheet.iter_rows(max_col=1, min_row=2, values_only=True):
            existing_orders.add(row[0])
        for order in data:
            if order[0] in existing_orders:
                duplicate_orders.append(order[0])
            else:
                worksheet.append(order)
        
        # if there are duplicate orders, show in a seperate window
        if duplicate_orders:
            GUI.show_duplicate_orders(duplicate_orders)

    # save
    def save(self, workbook: Workbook):
        """Saves the specified worksheet

        Args:
            worksheet (Worksheet): An instance of openpyxl.workbook.workbook.
        """
        try:
            workbook.save(self.__excel_filename)
        except PermissionError:
            # if the file is already opened by an editor
            message = {
                "title": "File in Use",
                "message": "The file 'boltworld.xlsx' is currently open.\n\n""Please close the Excel file and click OK to continue.",
                "icon": "warning"
            }
            # prompting the user for error message
            GUI.prompt_error(message=message)
            # saving workbook again. Can trigger two cases.
            #   1. Either the user has closed the Excel window
            #   2. The window is still open and the user clicked OK
            # saving the workbook again
            # For the first case, saving the workbook again will work fine, but
            # for the second case, we have to tell the GUI not to process the Excel file (ofcourse it won't and it'll generate another exceptoion
            # in which saving workbook denies the permission) and the process of writing order details will be dispersed.
            # For the second case, we make sure error messages in the frontend are correct and no false or corrupted changes has been done to file.
            try:
                workbook.save(self.__excel_filename)
            except PermissionError:
                return 101
        

        

