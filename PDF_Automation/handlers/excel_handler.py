
# This file contains the ExcelHandler class

# IMPORTS!
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from .gui_handler import GUI

class ExcelHandler:
    """This class is reponsible for handling excel related functionality such as reading, appending, removing, copying etc.
    """

    __excel_filename: str = ''              # Name of the excel file that is being managed

    # constructor
    def __init__(self, filename: str) -> None:
        """Initialize an ExcelHandler instance.
        """
        # Validating filename
        assert type(filename) == str, "Excel filename needs to be string"
        assert filename != "", "Excel filename cannot be none"
        
        # initializing
        self.__excel_filename = filename

    # open_file
    def open_file(self, headers: list, create_file: bool) -> openpyxl.Workbook:
        """Opens an Excel file, it also checks either the file aready exists or not,
        if not then it creates a new file with the specified headers.

        Args:
            filename (str): Name of the Excel file.
            headers (list): A list containing headers.

            Returns:
                _type_: None else openpyxl.Workbook instance.
        """
        wb = None
        # opening the excel file
        try:
            wb = openpyxl.load_workbook(filename=self.__excel_filename)
        except FileNotFoundError:       # if the file doesn't exist
            # if create_file is True, then proceed with creating the file
            if create_file:
                # creating an excel file
                wb = openpyxl.Workbook()
                # getting the active sheet
                ws = wb.active
                # appending headers to the top of the file
                ws.append(headers)
                # setting the title of the sheet
                ws.title = "Order_Details"
                # saving the workbook
                wb.save(self.__excel_filename)
            else:
                return 102
        
        return wb
    

    # search
    def search(self, _type: str, search_value: str, excel_filename: str):
        """Search for the order details from the given Excel file.
        """
        # opening the Excel file and don't create a new one 
        wb = self.open_file(headers=None, create_file=False)
        # searching 
        if wb == 102:
            # it means that the file didn't exist
            # need to prompt an error message to the user
            message = {
                "title": "No Data",
                "message": "The Excel file doesn't exist yet.\nProcess a PDF first to create the database."
            }
            GUI.prompt_error(code=wb, message=message)
            return (102, wb)   # will return 102 as its status code
        # if wb is OK
        results = []
        ws = wb.active
        find = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            # skipping empty rows, just in case
            if not row[0]:
                continue
            # now, searching on the basis of type
            if _type == "order":
                if str(row[0]) == search_value:
                    if not find:
                        find = True
                    results.append(row)
            elif _type == "date":
                if row[1] == search_value:
                    if not find:
                        find = True
                    results.append(row)
            elif _type == "user":
                if row[3] and search_value.lower() in str(row[3]).lower():
                    if not find:
                        find = True
                    results.append(row)
        if find:
            return (100, results)
        else:
            return (103, results)


    # write
    def write(self, worksheet: Worksheet, data):
        """Writes order details on the excel file.
        """
        # need to check for duplicate orders before adding
        duplicate_orders = []
        # fetch existing orders from excel
        existing_orders = set()
        for row in worksheet.iter_rows(max_col=1, min_row=2, values_only=True):
            existing_orders.add(row[0])
        for order in data:
            if order[0] in existing_orders:
                duplicate_orders.append(order[0])
            else:
                worksheet.append(order)
        
        # if there are duplicate orders, show in a seperate window
        if duplicate_orders:
            GUI.show_duplicate_orders(duplicate_orders)

    # save
    def save(self, workbook: Workbook):
        """Saves the specified worksheet

        Args:
            worksheet (Worksheet): An instance of openpyxl.workbook.workbook.
        """
        try:
            workbook.save(self.__excel_filename)
        except PermissionError:
            # if the file is already opened by an editor
            message = {
                "title": "File in Use",
                "message": "The file 'boltworld.xlsx' is currently open.\n\n""Please close the Excel file and click OK to continue.",
                "icon": "warning"
            }
            # prompting the user for error message
            GUI.prompt_error(code=101, message=message)
            # saving workbook again. Can trigger two cases.
            #   1. Either the user has closed the Excel window
            #   2. The window is still open and the user clicked OK
            # saving the workbook again
            # For the first case, saving the workbook again will work fine, but
            # for the second case, we have to tell the GUI not to process the Excel file (ofcourse it won't and it'll generate another exceptoion
            # in which saving workbook denies the permission) and the process of writing order details will be dispersed.
            # For the second case, we make sure error messages in the frontend are correct and no false or corrupted changes has been done to file.
            try:
                workbook.save(self.__excel_filename)
            except PermissionError:
                return 101
        

        

